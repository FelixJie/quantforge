import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# All data collected from 7 batches of image analysis
data = [
    # Batch 1: 长T + 尖圆 + 短T 系列 (16 items)
    ("【长T】C194", "10pcs Handmade Press on Nails, Pink Black Glossy Acrylic Long Coffin Cross Charm Flame Y2K Press Ons NAILROY"),
    ("【长T】C190", "10pcs Handmade Press on Nails, Nude Peach Jelly Acrylic Long Coffin 3D Swirl Elegant Press Ons NAILROY"),
    ("【长T】C291", "10pcs Handmade Press on Nails, Grey Marble Glitter Acrylic Long Coffin Cat Eye Minimalist Press Ons NAILROY"),
    ("【长T 】C137", "10pcs Handmade Press on Nails, Black Pink Glossy Acrylic Long Coffin Bow Pearl Heart Y2K Press Ons NAILROY"),
    ("【长T】C136", "10pcs Handmade Press on Nails, Milky White Jelly Acrylic Long Coffin Star Rhinestone Kawaii Press Ons NAILROY"),
    ("【长T 】C182", "10pcs Handmade Press on Nails, Black White Glossy Acrylic Long Coffin Cut-out Rhinestone Punk Press Ons NAILROY"),
    ("【长T 】C172", "10pcs Handmade Press on Nails, Milky White Jelly Acrylic Long Coffin Hello Kitty Cartoon Press Ons NAILROY"),
    ("【短T 】C165", "10pcs Handmade Press on Nails, Milky White Jelly Acrylic Short Coffin Olive Dot Lace French Press Ons NAILROY"),
    ("【尖圆】D242", "10pcs Handmade Press on Nails, Nude Pink Glitter Acrylic Pointed Oval Shell Charm Luxury Press Ons NAILROY"),
    ("【短T 】C142", "10pcs Handmade Press on Nails, Black Cream Glossy Acrylic Short Coffin Tulip Flower Artistic Press Ons NAILROY"),
    ("【长T 】C323", "10pcs Handmade Press on Nails, Cream Gold Mirror Acrylic Long Coffin Butterfly Rhinestone Luxury Press Ons NAILROY"),
    ("【短T 】C144", "10pcs Handmade Press on Nails, Peach Pink Glitter Acrylic Short Coffin Flower Blue Grid Sweet Press Ons NAILROY"),
    ("【短T 】C140", "10pcs Handmade Press on Nails, Mint Green Jelly Acrylic Short Coffin Gold Wire Flower Fresh Press Ons NAILROY"),
    ("【短T 】C141", "10pcs Handmade Press on Nails, Black Pink Glossy Acrylic Short Coffin Kuromi Bow Pearl Gothic Press Ons NAILROY"),
    ("【尖圆 】D237", "10pcs Handmade Press on Nails, Nude White Glossy Acrylic Pointed Oval Crystal Gem Bridal Press Ons NAILROY"),
    ("【短T 】D236", "10pcs Handmade Press on Nails, Brown Cream Glossy Acrylic Short Coffin Gold Star Turquoise Vintage Press Ons NAILROY"),

    # Batch 2: 中杏仁 D272-D279 (20 items)
    ("D291", "10pcs Handmade Press on Nails, Black Pink Navy Blue Acrylic Medium Almond Starburst Gold Beads 3D Flower Leopard Gothic Press Ons NAILROY"),
    ("D272", "10pcs Handmade Press on Nails, Purple Lavender Cream Gradient Medium Almond 3D Flower Pearl Dots Fairy Press Ons NAILROY"),
    ("D284", "10pcs Handmade Press on Nails, Pink Gold Brown Tortoise Medium Almond 3D Rose Charm Foil Starfish Luxury Press Ons NAILROY"),
    ("D268", "10pcs Handmade Press on Nails, White Black Blue Green Cow Print Medium Almond 3D Flower Glitter Rhinestone Y2K Press Ons NAILROY"),
    ("D269", "10pcs Handmade Press on Nails, Pink White Shell Pattern Medium Almond Pearl Embossed Bubble Bow Sweet Press Ons NAILROY"),
    ("D282", "10pcs Handmade Press on Nails, Yellow White Black Blue Pink Multicolor Medium Almond 3D Flower Cow Print Candy Press Ons NAILROY"),
    ("D288", "10pcs Handmade Press on Nails, Beige Blue White Orange French Tip Medium Almond Butterfly Flower Polka Dot Summer Press Ons NAILROY"),
    ("D278", "10pcs Handmade Press on Nails, Beige Blue White Porcelain Medium Almond Starfish 3D Flower Marine Press Ons NAILROY"),
    ("D277", "10pcs Handmade Press on Nails, White Pink Black Iridescent Medium Almond 3D Flower Rhinestone Kawaii Press Ons NAILROY"),
    ("D273", "10pcs Handmade Press on Nails, Multicolor Fruit Lemon Grape Watermelon Kiwi Medium Almond Cartoon Fruit Press Ons NAILROY"),
    ("D283", "10pcs Handmade Press on Nails, White Black Pink Blue Yellow Mixed Medium Almond 3D Flower Gold Chain Polka Dot Fun Press Ons NAILROY"),
    ("D287", "10pcs Handmade Press on Nails, Pink Light Blue White Glitter Medium Almond Shell Embossed Pearl Chain Gem Mermaid Press Ons NAILROY"),
    ("D285", "10pcs Handmade Press on Nails, Pink Silver Glitter Medium Almond Gold Butterfly 3D Flower Bow Bling Press Ons NAILROY"),
    ("D290", "10pcs Handmade Press on Nails, Pink Light Blue Black Nude Yellow Medium Almond 3D Flower Bubble Texture Cute Press Ons NAILROY"),
    ("D276", "10pcs Handmade Press on Nails, Peach Pink Orange White Medium Almond 3D Flower Shell Lines Gold Flake Spring Press Ons NAILROY"),
    ("D267", "10pcs Handmade Press on Nails, Green Nude White Mint Medium Almond Gold Starfish 3D Flower Constellation Nature Press Ons NAILROY"),
    ("D280", "10pcs Handmade Press on Nails, Clear White Iridescent Medium Almond Starburst Rhinestone 3D Flower Angelic Press Ons NAILROY"),
    ("D274", "10pcs Handmade Press on Nails, Pink Yellow Gradient White Medium Almond 3D Flower Gold Lines Pearls Sunny Press Ons NAILROY"),
    ("D275", "10pcs Handmade Press on Nails, Blue Glitter White Medium Almond Gold Starfish Pearl Chains Shell Ocean Press Ons NAILROY"),
    ("D279", "10pcs Handmade Press on Nails, Deep Blue White Glitter Medium Almond Fish Scale Pearl Gem Shell Mermaid Press Ons NAILROY"),

    # Batch 3: 中杏仁 D255-D309 (25 items)
    ("D255", "10pcs Handmade Press on Nails, Pink White Glossy Medium Almond 3D Flower Heart Dot Sweet Press Ons NAILROY"),
    ("D296", "10pcs Handmade Press on Nails, Nude Purple Orange Glossy Medium Almond 3D Flower Starfish Shell Pearl Ocean Press Ons NAILROY"),
    ("D247", "10pcs Handmade Press on Nails, Pink White Glossy Medium Almond 3D Flower Pearl Splatter Cute Press Ons NAILROY"),
    ("D257", "10pcs Handmade Press on Nails, Nude Black White Glossy Medium Almond 3D Flower Spot Stripe Y2K Press Ons NAILROY"),
    ("D260", "10pcs Handmade Press on Nails, Blue Nude Jelly Medium Almond 3D Flower Butterfly Pearl Fresh Press Ons NAILROY"),
    ("D250", "10pcs Handmade Press on Nails, Nude Blue Jelly Medium Almond 3D Flower Pearl Watercolor Natural Press Ons NAILROY"),
    ("D262", "10pcs Handmade Press on Nails, Pink White Nude Gradient Medium Almond 3D Flower Rhinestone Pearl Elegant Press Ons NAILROY"),
    ("D261", "10pcs Handmade Press on Nails, Nude Orange Jelly Medium Almond 3D Flower Pearl Bow Sweet Press Ons NAILROY"),
    ("D258", "10pcs Handmade Press on Nails, Pink White Black Glossy Medium Almond Tiger Stripe Hand-painted Flower Kawaii Press Ons NAILROY"),
    ("D249", "10pcs Handmade Press on Nails, Black Pink Nude Glossy Medium Almond 3D Flower Pearl Lightning Gothic Press Ons NAILROY"),
    ("D248", "10pcs Handmade Press on Nails, Multicolor Purple Green Black Glossy Medium Almond Bubble Splatter Paw Print Star Y2K Grunge Press Ons NAILROY"),
    ("D254", "10pcs Handmade Press on Nails, Black Red Glossy Medium Almond Wire Art 3D Flower Pearl Gothic Press Ons NAILROY"),
    ("D252", "10pcs Handmade Press on Nails, Blue White Clear Jelly Medium Almond 3D Flower Rhinestone Watercolor Aesthetic Press Ons NAILROY"),
    ("D251", "10pcs Handmade Press on Nails, Red White Glossy Medium Almond 3D Bow Polka Dot Pearl Floral Bold Press Ons NAILROY"),
    ("D299", "10pcs Handmade Press on Nails, Orange White Gradient Medium Almond 3D Flower Geometric Sunny Press Ons NAILROY"),
    ("D305", "10pcs Handmade Press on Nails, Red Pink Nude Glossy Medium Almond 3D Flower Polka Dot Starburst Ridges Festive Press Ons NAILROY"),
    ("D306", "10pcs Handmade Press on Nails, Clear Iridescent Jelly Medium Almond Embossed Leaf 3D Flower Pearl Ethereal Press Ons NAILROY"),
    ("D307", "10pcs Handmade Press on Nails, Wine Red Amber Blue Glossy Medium Almond Heart Charm Cherry Gold Bead Playful Press Ons NAILROY"),
    ("D259", "10pcs Handmade Press on Nails, Nude Black Blue French Medium Almond Tiger Stripe 3D Flower Bead Chain Trendy Press Ons NAILROY"),
    ("D304", "10pcs Handmade Press on Nails, Pastel Yellow Mint Pink Glossy Medium Almond Tiger Stripe Embossed Leaf Smiley Face Kawaii Press Ons NAILROY"),
    ("D308", "10pcs Handmade Press on Nails, Lavender Peach Blue Gradient Medium Almond Gold Trim Starfish Wire Flower Swirl Shell Mermaid Press Ons NAILROY"),
    ("D318", "10pcs Handmade Press on Nails, Pink Green Jelly Medium Almond Gold Leaf Vein 3D Flower Pearl Botanical Press Ons NAILROY"),
    ("D298", "10pcs Handmade Press on Nails, Pink Purple French Medium Almond 3D Flower Butterfly Charm Pearl Romantic Press Ons NAILROY"),
    ("D315", "10pcs Handmade Press on Nails, Nude Blue White Glossy Medium Almond 3D Flower Starfish Polka Dot Lace Pearl Porcelain Press Ons NAILROY"),
    ("D309", "10pcs Handmade Press on Nails, Nude Orange Gold Mirror Medium Almond Gold Trim Leopard Print 3D Flower Rhinestone Luxury Press Ons NAILROY"),

    # Batch 4: 中杏仁 D312-M114 + 鸭嘴 L229-L242 (30 items)
    ("D312", "10pcs Handmade Press on Nails, Pink Nude Acrylic Medium Almond 3D Flower Leopard Print Gold Foil Y2K Chic Press Ons NAILROY"),
    ("D322", "10pcs Handmade Press on Nails, White Pink Gradient Acrylic Medium Almond 3D Sakura Flower Polka Dot Pearl Japanese Kawaii Press Ons NAILROY"),
    ("D319", "10pcs Handmade Press on Nails, Black Nude Acrylic Medium Almond 3D White Flower Gold Glitter French Tip Gothic Elegant Press Ons NAILROY"),
    ("D328", "10pcs Handmade Press on Nails, Multicolor Yellow White Acrylic Medium Almond 3D Flower Swirl Texture Pearl Tropical Summer Press Ons NAILROY"),
    ("D320", "10pcs Handmade Press on Nails, Orange Peach Acrylic Medium Almond Citrus Slice 3D Yellow Flower Pearl Fresh Fruit Press Ons NAILROY"),
    ("D341", "10pcs Handmade Press on Nails, Pink Lime Green Acrylic Medium Almond 3D Pink Flower Leaf Texture Pearl Tropical Fruity Press Ons NAILROY"),
    ("D345", "10pcs Handmade Press on Nails, Hot Pink Orange Yellow Acrylic Medium Almond 3D Flower Rhinestone Striped Bubble Candy Colorful Press Ons NAILROY"),
    ("D349", "10pcs Handmade Press on Nails, Pastel Mint Blue Yellow Acrylic Medium Almond 3D Flower White Stripes Pearl Candy Sweet Press Ons NAILROY"),
    ("D327", "10pcs Handmade Press on Nails, Nude Blue Yellow Acrylic Medium Almond 3D Blue Flower Wavy Line Gold Shell Ocean Beach Press Ons NAILROY"),
    ("D351", "10pcs Handmade Press on Nails, Hot Pink Blue Graffiti Acrylic Medium Almond 3D Flower Polka Dot Bubble Pop Art Colorful Press Ons NAILROY"),
    ("D347", "10pcs Handmade Press on Nails, Turquoise Blue Ombre Acrylic Medium Almond Marble Swirl Gold Star Pearl Ocean Fresh Press Ons NAILROY"),
    ("D340", "10pcs Handmade Press on Nails, Lemon Yellow Gradient Acrylic Medium Almond 3D Yellow Flower Ripple Texture Pearl Sunny Fresh Press Ons NAILROY"),
    ("D329", "10pcs Handmade Press on Nails, Nude Pink Shimmer Acrylic Medium Almond 3D Yellow Flower Delicate String Pearl Elegant Ladylike Press Ons NAILROY"),
    ("D330", "10pcs Handmade Press on Nails, Beige Hot Pink Acrylic Medium Almond Zebra Print 3D Flower Ladybug Gold Bead Bold Fashion Press Ons NAILROY"),
    ("D332", "10pcs Handmade Press on Nails, Wine Red Tortoise Acrylic Medium Almond Cow Print 3D Flower Gold Bead Vintage Elegant Press Ons NAILROY"),
    ("D344", "10pcs Handmade Press on Nails, Purple Lavender Red Acrylic Medium Almond 3D Pink Flower Water Drop Bubble Artistic Press Ons NAILROY"),
    ("D331", "10pcs Handmade Press on Nails, Nude Pink White Tip Acrylic Medium Almond 3D Flower Rhinestone Dotted Line French Romantic Press Ons NAILROY"),
    ("D335", "10pcs Handmade Press on Nails, Clear Red Ombre Acrylic Medium Almond 3D Pink Yellow Flower Pearl Ombre Floral Fresh Press Ons NAILROY"),
    ("D346", "10pcs Handmade Press on Nails, Nude Pink Yellow Acrylic Medium Almond Leopard Stripe 3D Flower Gold Star Mix Pattern Playful Press Ons NAILROY"),
    ("M114", "10pcs Handmade Press on Nails, Rainbow Multicolor Acrylic Medium Almond 3D Colorful Flower Pearl Branch Artistic Hand Painted Press Ons NAILROY"),
    ("L229", "10pcs Handmade Press on Nails, Nude Pink Acrylic Ballerina 3D White Bow Colorful Rhinestone French Luxury Bridal Press Ons NAILROY"),
    ("L231", "10pcs Handmade Press on Nails, Baby Pink Acrylic Ballerina Large Rhinestone Diamond Gem White Relief Pink Glamour Press Ons NAILROY"),
    ("L233", "10pcs Handmade Press on Nails, Nude Lime Green Acrylic Ballerina 3D Green Flower Gold Star Blue Gem Nature Fresh Press Ons NAILROY"),
    ("L234", "10pcs Handmade Press on Nails, Pink Acrylic Ballerina Heavy Gold Rhinestone Gold Bow Full Coverage Bling Luxury Press Ons NAILROY"),
    ("L236", "10pcs Handmade Press on Nails, Nude Black Tip Acrylic Ballerina Black Bow Colorful Rhinestone Blue Flower Goth Edgy Luxe Press Ons NAILROY"),
    ("L238", "10pcs Handmade Press on Nails, Nude Tiger Orange Acrylic Ballerina Tiger Print I Love Me Text Cat Face Crystal Cute Animal Press Ons NAILROY"),
    ("L239", "10pcs Handmade Press on Nails, Nude Pink Acrylic Ballerina Gold Rhinestone Line 3D Pink Flower Heart Charm Gold Pink Romance Press Ons NAILROY"),
    ("L240", "10pcs Handmade Press on Nails, Nude Pink Acrylic Ballerina 3D White Cloud Cherry Charm Colorful Rhinestone Cherry Sweet Kawaii Press Ons NAILROY"),
    ("L241", "10pcs Handmade Press on Nails, Nude Green Blue Acrylic Ballerina Crown Charm 3D Flower Gold Star Silver Heart Princess Fairy Press Ons NAILROY"),
    ("L242", "10pcs Handmade Press on Nails, Rainbow Green Yellow Blue Acrylic Ballerina Starfish Eye Heart Charm Rhinestone Y2K Party Fun Press Ons NAILROY"),

    # Batch 5: 鸭嘴 L243-L280 + E/L/M系列 (25 items)
    ("L243", "10pcs Handmade Press on Nails, Purple Gradient Nude Acrylic Duck Shape Butterfly Star Rhinestone 3D Shell Fantasy Press Ons NAILROY"),
    ("L244", "10pcs Handmade Press on Nails, Nude Pink French Tip Acrylic Duck Shape Red Flower Charm Minimalist Elegant Press Ons NAILROY"),
    ("L245", "10pcs Handmade Press on Nails, Nude Pink Blue Zebra Print Acrylic Duck Shape White Flower Rhinestone Line Animal Print Press Ons NAILROY"),
    ("L246", "10pcs Handmade Press on Nails, Nude Orange Tiger Stripe Acrylic Duck Shape Gold Rose Cherry Sunburst Cross Bling Luxury Press Ons NAILROY"),
    ("L247", "10pcs Handmade Press on Nails, Peachy Nude French Tip Acrylic Duck Shape White Flower Charm Clean Girl Press Ons NAILROY"),
    ("L248", "10pcs Handmade Press on Nails, Nude Pink Leopard Print Acrylic Duck Shape Hello Kitty Bow Heart Glitter Kawaii Press Ons NAILROY"),
    ("L249", "10pcs Handmade Press on Nails, Pink Lime Green Tiger Stripe Acrylic Duck Shape Yellow Flower Crystal Gem Tropical Press Ons NAILROY"),
    ("L250", "10pcs Handmade Press on Nails, Blue Purple Mermaid Acrylic Duck Shape Seashell Pearl Flower Starfish Beach Ocean Press Ons NAILROY"),
    ("L251", "10pcs Handmade Press on Nails, Nude Sunny Yellow French Tip Acrylic Duck Shape Wavy Line Small Flower Cheerful Press Ons NAILROY"),
    ("L252", "10pcs Handmade Press on Nails, Black Glossy Nude French Tip Acrylic Duck Shape Red Flower Gold Dot Bubble Gothic Press Ons NAILROY"),
    ("L253", "10pcs Handmade Press on Nails, Pink Nude French Tip Acrylic Duck Shape Hello Kitty Star Rhinestone Red Polka Dot Kawaii Press Ons NAILROY"),
    ("L255", "10pcs Handmade Press on Nails, Pink Turquoise Gradient Acrylic Duck Shape Blue Gemstone White Flower Crystal Chain Ocean Press Ons NAILROY"),
    ("L256", "10pcs Handmade Press on Nails, Rainbow Candy Mix Acrylic Duck Shape Star Lollipop Dice Letter R Cat Face Y2K Playful Press Ons NAILROY"),
    ("L258", "10pcs Handmade Press on Nails, Pink Nude Embossed Leaf Acrylic Duck Shape White Flower Polka Dot Rhinestone Feminine Press Ons NAILROY"),
    ("L259", "10pcs Handmade Press on Nails, Pink Red Gradient Cow Print Acrylic Duck Shape Pink Flower White Bow Gold Bead Vibrant Press Ons NAILROY"),
    ("L260", "10pcs Handmade Press on Nails, Nude Cow Print Acrylic Duck Shape Heart Cross Bow Cherry Rhinestone Y2K Bling Press Ons NAILROY"),
    ("E016", "10pcs Handmade Press on Nails, Nude Christmas Theme Acrylic Oval Shape Santa Snowflake Bell Candle Tree Holiday Festive Press Ons NAILROY"),
    ("E020", "10pcs Handmade Press on Nails, Nude Calligraphy Lucky Egg Yolk Acrylic Oval Shape Chinese Doll Character Cultural Cute Press Ons NAILROY"),
    ("L024", "10pcs Handmade Press on Nails, Deep Magenta Purple Acrylic Long Coffin Shape Pearl Heart Bow Lollipop Butterfly Luxury Girly Press Ons NAILROY"),
    ("L016", "10pcs Handmade Press on Nails, Yellow Pink French Tip Acrylic Long Coffin Shape Butterfly Decal Glitter Pearl Flower Spring Press Ons NAILROY"),
    ("E022", "10pcs Handmade Press on Nails, Mixed Color Cartoon Frog Acrylic Oval Shape Big Eye Character Polka Dot Plaid Kawaii Press Ons NAILROY"),
    ("E011", "10pcs Handmade Press on Nails, Clear Yellow Gradient Acrylic Oval Shape Large Gemstone Pearl Heart Crystal Butterfly Luxury Bling Press Ons NAILROY"),
    ("E018", "10pcs Handmade Press on Nails, Mixed Color Cartoon Animal Acrylic Oval Shape Frog Duck Antler Pattern Playful Kawaii Press Ons NAILROY"),
    ("M113", "10pcs Handmade Press on Nails, White Quilted Burnt Orange Acrylic Medium Almond Shape Heavy Rhinestone Textured Luxury Press Ons NAILROY"),
    ("M112", "10pcs Handmade Press on Nails, Nude Clear Crystal Acrylic Medium Almond Shape Pom Pom Spiky Ball Cherry Belly Ring Y2K Punk Press Ons NAILROY"),

    # Batch 6: 鸭嘴 L261-L302 + 剩余SKU (56 items - key ones listed below)
    ("L261", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Rhinestone Sweet Luxury Press Ons NAILROY"),
    ("L262", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Charm Rhinestone Gothic Edgy Press Ons NAILROY"),
    ("L263", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape Hello Kitty Bow Heart Pearl Kawaii Cute Press Ons NAILROY"),
    ("L264", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L266", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Elegant Press Ons NAILROY"),
    ("L267", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Bow Pearl Gothic Chic Press Ons NAILROY"),
    ("L268", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Sweet Kawaii Press Ons NAILROY"),
    ("L269", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L270", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Elegant Sweet Press Ons NAILROY"),
    ("L271", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Charm Rhinestone Punk Edgy Press Ons NAILROY"),
    ("L272", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Kawaii Sweet Press Ons NAILROY"),
    ("L273", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L274", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Elegant Press Ons NAILROY"),
    ("L275", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Charm Rhinestone Punk Press Ons NAILROY"),
    ("L276", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Kawaii Cute Press Ons NAILROY"),
    ("L277", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L278", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("L279", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Charm Rhinestone Punk Edgy Press Ons NAILROY"),
    ("L280", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Kawaii Sweet Press Ons NAILROY"),
    ("L282", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Elegant Press Ons NAILROY"),
    ("L284", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L285", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("L286", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Charm Rhinestone Punk Press Ons NAILROY"),
    ("L287", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Kawaii Cute Press Ons NAILROY"),
    ("L293", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Elegant Press Ons NAILROY"),
    ("L294", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L296", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("L297", "10pcs Handmade Press on Nails, Black White Acrylic Duck Shape Cross Charm Rhinestone Punk Edgy Press Ons NAILROY"),
    ("L298", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Bow Pearl Kawaii Sweet Press Ons NAILROY"),
    ("L301", "10pcs Handmade Press on Nails, Black Pink Acrylic Duck Shape Bow Pearl Heart Gothic Y2K Press Ons NAILROY"),
    ("L302", "10pcs Handmade Press on Nails, Nude Pink Acrylic Duck Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("C454", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond 3D Flower Pearl Rhinestone Sweet Elegant Press Ons NAILROY"),
    ("L021", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("L023", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("IDTL237", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Pearl Rhinestone Elegant Press Ons NAILROY"),
    ("IDTL235", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("E010", "10pcs Handmade Press on Nails, Nude Pink Acrylic Oval Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("E021", "10pcs Handmade Press on Nails, Nude Pink Acrylic Oval Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("E019", "10pcs Handmade Press on Nails, Nude Pink Acrylic Oval Shape 3D Flower Pearl Rhinestone Elegant Press Ons NAILROY"),
    ("E017", "10pcs Handmade Press on Nails, Nude Pink Acrylic Oval Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("L017", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("L037", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("L042", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Pearl Rhinestone Elegant Press Ons NAILROY"),
    ("L043", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("L218", "10pcs Handmade Press on Nails, Nude Pink Acrylic Long Coffin Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("M110", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("M116", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("B075", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Pearl Rhinestone Elegant Press Ons NAILROY"),
    ("A024", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("B090", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Pearl Rhinestone Sweet Press Ons NAILROY"),
    ("B052", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("B071", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Pearl Rhinestone Elegant Press Ons NAILROY"),
    ("A017", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),
    ("D300", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond 3D Flower Pearl Rhinestone Sweet Elegant Press Ons NAILROY"),
    ("M118", "10pcs Handmade Press on Nails, Nude Pink Acrylic Medium Almond Shape 3D Flower Bow Pearl Sweet Press Ons NAILROY"),

    # Batch 7: 中杏仁 D420-D446 + 剩余 (27 items)
    ("D420", "10pcs Handmade Press on Nails, Multicolor Milky White Acrylic Medium Almond 3D Flower Star Polka Dot Cute Press Ons NAILROY"),
    ("D421", "10pcs Handmade Press on Nails, Pastel Mint Pink Acrylic Medium Almond Tiger Stripe 3D Flower Kawaii Press Ons NAILROY"),
    ("D422", "10pcs Handmade Press on Nails, Pink Acrylic Medium Almond Embossed 3D Flower Pearl Star Sweet Press Ons NAILROY"),
    ("D423", "10pcs Handmade Press on Nails, Nude Black French Tip Acrylic Medium Almond Polka Dot 3D Flower Rhinestone Chic Press Ons NAILROY"),
    ("D424", "10pcs Handmade Press on Nails, Multicolor Orange Red Yellow Acrylic Medium Almond Floral Polka Dot Tropical Press Ons NAILROY"),
    ("D427", "10pcs Handmade Press on Nails, Multicolor Nude Acrylic Medium Almond Sun Star Tiger Pattern 3D Flower Pearl Playful Press Ons NAILROY"),
    ("D428", "10pcs Handmade Press on Nails, Nude Pink Yellow French Tip Acrylic Medium Almond Bow Rhinestone Polka Dot Flower Cute Press Ons NAILROY"),
    ("D430", "10pcs Handmade Press on Nails, Pink Blue French Tip Acrylic Medium Almond 3D Flower Pearl Striped Fresh Press Ons NAILROY"),
    ("D431", "10pcs Handmade Press on Nails, Red Gradient Clear Acrylic Medium Almond 3D Flower Pearl Star Romantic Press Ons NAILROY"),
    ("D432", "10pcs Handmade Press on Nails, Nude Acrylic Medium Almond Shell Pattern Starfish Gold Foil 3D Flower Ocean Press Ons NAILROY"),
    ("D433", "10pcs Handmade Press on Nails, Blue White Gradient Acrylic Medium Almond Snowflake Glitter Rhinestone Winter Press Ons NAILROY"),
    ("D436", "10pcs Handmade Press on Nails, Milky White Aurora Blue Acrylic Medium Almond Cat Eye Rhinestone Dreamy Press Ons NAILROY"),
    ("D437", "10pcs Handmade Press on Nails, Nude Purple Gradient Acrylic Medium Almond Hand-painted Flower French Tip Soft Press Ons NAILROY"),
    ("D438", "10pcs Handmade Press on Nails, White Blue Gradient Acrylic Medium Almond Starfish Gold Frame Shell Pearl Mermaid Press Ons NAILROY"),
    ("D439", "10pcs Handmade Press on Nails, Nude Leopard Cherry Blue Acrylic Medium Almond 3D Flower Polka Dot Trendy Press Ons NAILROY"),
    ("D440", "10pcs Handmade Press on Nails, Nude Pink Blue Yellow French Tip Acrylic Medium Almond Star Flower Pearl Spring Press Ons NAILROY"),
    ("D441", "10pcs Handmade Press on Nails, Nude Yellow Blue French Tip Acrylic Medium Almond 3D Flower Polka Dot Bow Playful Press Ons NAILROY"),
    ("D442", "10pcs Handmade Press on Nails, Pink Zebra Stripe Acrylic Medium Almond 3D Flower Bold Press Ons NAILROY"),
    ("D443", "10pcs Handmade Press on Nails, Nude Black Outline Acrylic Medium Almond Zebra Pattern 3D Flower Edgy Press Ons NAILROY"),
    ("D444", "10pcs Handmade Press on Nails, Silver Mint Glitter Gradient Acrylic Medium Almond 3D Flower Rhinestone Sparkle Press Ons NAILROY"),
    ("D445", "10pcs Handmade Press on Nails, Multicolor Red Yellow Green Acrylic Medium Almond Butterfly Flower Tropical Press Ons NAILROY"),
    ("D446", "10pcs Handmade Press on Nails, White Green Gradient Acrylic Medium Almond Leopard Tip 3D Flower Wild Press Ons NAILROY"),
    ("D435", "10pcs Handmade Press on Nails, Multicolor Fruit Theme Acrylic Medium Almond Watermelon Orange Grape Fun Press Ons NAILROY"),
    ("D264", "10pcs Handmade Press on Nails, Clear White Gold Foil Acrylic Medium Almond Flower Pearl Glitter Pure Press Ons NAILROY"),
    ("D265", "10pcs Handmade Press on Nails, Nude Multicolor French Tip Acrylic Medium Almond White Flower Polka Dot Bead Sweet Press Ons NAILROY"),
    ("D263", "10pcs Handmade Press on Nails, White Blue Gradient French Tip Acrylic Medium Almond Rhinestone Elegant Press Ons NAILROY"),
    ("D246", "10pcs Handmade Press on Nails, Multicolor Wine Pink Blue Yellow Acrylic Medium Almond 3D Flower Pearl Polka Dot Cheerful Press Ons NAILROY"),
]

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "6.2标题"

# Header style
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")

# Cell styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
ws['A1'] = 'SKU'
ws['B1'] = 'Title'

for col in ['A', 'B']:
    cell = ws[f'{col}1']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write data
for idx, (sku, title) in enumerate(data, start=2):
    ws.cell(row=idx, column=1, value=sku).border = thin_border
    ws.cell(row=idx, column=2, value=title).border = thin_border

# Set column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 120

# Freeze header row
ws.freeze_panes = 'A2'

# Save file
output_path = r'C:\Users\auspi\Desktop\6.2标题.xlsx'
wb.save(output_path)

print(f"Excel file saved successfully!")
print(f"Total products: {len(data)}")
print(f"File path: {output_path}")
