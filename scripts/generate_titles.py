# -*- coding: utf-8 -*-
"""
智能美甲产品标题生成器
基于SKU名称规则 + 多样化颜色/材质/装饰词库，为每个SKU生成唯一高质量英文标题
"""
import openpyxl
import re
import hashlib

# ============ 词库 ============
COLORS = [
    "Nude", "Blush Pink", "Milky White", "Baby Blue", "Lavender", "Hot Pink",
    "Dusty Rose", "Cherry Red", "Burgundy", "Wine Red", "Mint Green", "Sage Green",
    "Olive", "Peach", "Coral", "Champagne Gold", "Silver Chrome", "Rose Gold",
    "Black", "Ivory", "Pearl White", "Mocha", "Caramel", "Turquoise", "Teal",
    "Periwinkle", "Plum", "Ruby", "Fuchsia", "Emerald", "Honey", "Apricot",
    "Cinnamon", "Magenta Purple", "Royal Blue", "Electric Pink", "Lime Green",
    "Peach Orange", "Cream Yellow", "Coffee Brown", "Gray Violet", "Dark Navy"
]

MATERIALS = [
    "Glossy", "Matte", "Chrome", "Cat Eye", "Jelly Glazed", "Pearl",
    "Glitter", "French Tip", "Ombre", "Shimmer", "Metallic", "Iridescent",
    "Holographic", " Aurora Borealis", "Sugar", "Frosted", "Mirror", "Gradient"
]

DECORATIONS = [
    "Simple Minimalist", "Solid Color", "Rhinestone", "Crystal", "Pearl",
    "Butterfly", "Flower", "Bow", "Heart", "Star", "Moon", "AB Diamond",
    "Gold Foil", "Silver Wire", "Gem Stone", "Charms", "Bling Luxury",
    "Hand Painted", "3D Decor", "Sparkle", "Elegant Bridal", "Kawaii Cute",
    "Gothic Edgy", "Vintage Floral", "Geometric Pattern", "Animal Print",
    "Marble Swirl", "Tie Dye", "Zebra Stripe", "Leopard Print", "Cow Print",
    "Camouflage", "Checkered Plaid", "Polka Dot", "Striped Line", "Abstract Art",
    "Watercolor", "Aurora Shimmer", "Moonlight Glow", "Sunset Gradient",
    "Ocean Wave", "Galaxy Star", "Cloud Dream", "Candy Sweet", "Fruit Fresh",
    "Lolita Bow", "Princess Crown", "Angel Wing", "Devil Horn", "Skull Punk",
    "Chain Metal", "Feather Light", "Lace Elegant", "Velvet Texture",
    "Snake Skin", "Tortoise Shell", "Wood Grain", "Denim Jeans"
]

SHAPES_MAP = {
    # 短款系列
    "短款C": "Short Coffin", "短款D": "Short Coffin", "短款E": "Short Coffin",
    "短款L": "Short Almond", "短鸭嘴": "Short Ballerina",
    "短水管": "Short Coffin", "短T": "Short Round",
    # 长款系列
    "长款": "Long Coffin", "长款A": "Long Coffin", "长款B": "Long Coffin",
    "长款C": "Long Coffin", "长款D": "Long Coffin",
    "长款JSA": "Long Stiletto", "长款JSB": "Long Stiletto",
    "长款K": "Long Coffin", "长款KK": "Long Coffin",
    "长款L": "Long Almond", "长款M": "Long Almond",
    "长款N": "Long Almond", "长款O": "Long Almond",
    "长款O-": "Long Almond",
    "长蝴蝶结": "Long Almond", "长水管": "Long Coffin",
    "长TC": "Long Coffin", "长尖": "Long Stiletto",
    "长尖款": "Long Stiletto", "长方款": "Long Square",
    "长杏仁": "Long Almond",
    # 中款系列
    "中款": "Medium Almond", "中仁杏": "Medium Almond",
    "中杏仁": "Medium Almond", "中杳仁": "Medium Almond",
    "中长款": "Medium Long",
    # 杏仁款
    "杏仁款": "Almond", "杏仁甲": "Almond", "杏仁": "Almond",
    # 其他形状
    "鸭嘴款": "Ballerina", "穿戴式": "Stiletto",
    "细长尖": "Stiletto", "雕花": "Carved Almond",
    "尖圆": "Round Almond", "尖长款": "Long Round",
    "主图款": "Almond",
}


def get_shape(sku: str) -> str:
    """根据SKU名判断形状"""
    # 按长度降序匹配，避免短前缀误匹配
    for prefix, shape in sorted(SHAPES_MAP.items(), key=lambda x: -len(x[0])):
        if sku.startswith(prefix):
            return shape
    return "Almond"  # 默认


def sku_to_seed(sku: str) -> int:
    """SKU名转确定性的种子值"""
    return int(hashlib.md5(sku.encode()).hexdigest()[:8], 16)


def pick_from_list(lst: list, seed: int, index: int) -> str:
    """从词库中选择"""
    return lst[(seed + index * 7 + index * index * 13) % len(lst)]


def generate_title(sku: str) -> str:
    """为单个SKU生成唯一标题"""
    seed = sku_to_seed(sku)
    shape = get_shape(sku)

    # 选择颜色（1-2个）
    color1 = pick_from_list(COLORS, seed, 0)
    color2_option = pick_from_list(COLORS, seed, 1)

    # 选择材质（确保不同）
    mat1 = pick_from_list(MATERIALS, seed, 2)
    mat2_idx = (seed + 50) % len(MATERIALS)  # 偏移确保不同
    mat2_option = MATERIALS[mat2_idx] if MATERIALS[mat2_idx] != mat1 else pick_from_list(MATERIALS, seed+99, 3)

    # 选择装饰
    decor = pick_from_list(DECORATIONS, seed, 4)

    # 根据seed决定组合模式（0-9）
    mode = seed % 10

    if mode <= 1:
        # 简洁模式：颜色+材质+形状
        desc = f"{color1} {mat1}"
    elif mode <= 4:
        # 标准模式：颜色+材质+装饰+形状
        desc = f"{color1} {mat1} {decor} {shape}"
    elif mode <= 6:
        # 渐变/双色模式
        desc = f"{color1} {color2_option} Ombre {decor} {shape}"
    elif mode <= 8:
        # 奢华模式：颜色+双材质+装饰+形状
        desc = f"{color1} {mat1} {mat2_option} {decor} {shape}"
    else:
        # 特殊艺术模式
        desc = f"{color1} Hand Painted {decor} {shape}"

    title = f"10pcs Handmade Press on Nails, {desc} Y2K Coquette Bridal Prom Graduation Nailrosy"
    return title


def main():
    excel_path = r"C:\Users\auspi\Desktop\6.2标题.xlsx"

    # 读取现有Excel
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    placeholder_patterns = ["Nude Glossy Almond", "Nude Glossy Short Coffin"]
    # Also detect any title starting with Nude Glossy (old batch artifacts)
    # Also detect double materials
    import re
    def needs_regenerate(title):
        if any(p in title for p in placeholder_patterns):
            return True
        if title.startswith("10pcs Handmade Press on Nails, Nude Glossy"):
            return True
        if any(f'{m} {m}' in title for m in ['Glitter','Glossy','Matte','Chrome']):
            return True
        return False

    updated_count = 0
    total_placeholder = 0

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    new_data = []

    for row in rows:
        if not row[0]:
            continue
        sku = row[0]
        old_title = str(row[1]) if row[1] else ""

        if needs_regenerate(old_title):
            total_placeholder += 1
            new_title = generate_title(sku)
            new_data.append((sku, new_title))
            updated_count += 1
        else:
            new_data.append((sku, old_title))

    # 写回Excel
    for i, (sku, title) in enumerate(new_data, start=2):
        ws.cell(row=i, column=1, value=sku)
        ws.cell(row=i, column=2, value=title)

    wb.save(excel_path)

    print(f"Done! Updated {updated_count}/{total_placeholder} placeholder titles")
    print(f"Total rows: {len(new_data)}")


if __name__ == "__main__":
    main()
